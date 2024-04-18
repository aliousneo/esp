from __future__ import annotations
import re
import json
import pandas as pd
from typing import ClassVar
from elasticsearch import AsyncElasticsearch
from elasticsearch.helpers import async_streaming_bulk
from openai import OpenAI
import string

open_ai_client = OpenAI(api_key="sk-proj-e1s0JQ28eyqAstl3Dn1LT3BlbkFJ7mALdibiQ5jy7GYiLCUv")

ES_INDEX_NAME = "vector_search"
# ES_INDEX_NAME = "product_search_with_dimensions_updated"


def get_embedding(text: str, model: str = "text-embedding-ada-002"):
    text = text.replace("\n", " ")
    return open_ai_client.embeddings.create(input=[text], model=model).data[0].embedding


class AsyncElasticService:

    client: ClassVar[AsyncElasticsearch|None] = None
    products_info_json_filename: ClassVar[str] = 'all_data.json'
    products_info_csv_filename: ClassVar[str] = 'all_data.csv'
    inbound_data_excel_files: ClassVar[dict[str, str]] = {
        'products_descriptions': '/Users/aleksandrtynanov/Documents/work/other/esp/ProductList 2024.xlsx',
        'ai_search_patterns': '/Users/aleksandrtynanov/Documents/work/other/esp/Customer Data for AI Searching.xlsx'
    }

    es_index_name: ClassVar[str] = ES_INDEX_NAME
    products_es_model: ClassVar[dict[str, dict[str, dict[str, str|int|bool]]]] = {
        'properties': {
            'product_key': {
                'type': 'text'
            },
            'description': {
                'type': 'text'
            },
            'search_patterns': {
                'type': 'text'
            },
            "dimensions": {
                "type": 'integer'
            },
            "main_description_part": {
                'type': 'text'
            },
            "normalized_description": {
                'type': 'text'
            },
            "normalized_search_patterns": {
                'type': 'text'
            },
            'search_vector': {
                "type": "sparse_vector"
            }
        }
    }
    product_codes_patterns = [
        r'^[A-Z]{2}-\d{3}-\d{2,3}\s?[A-Z]{0,2}$',

    ]

    # 34b839768db6442c8ffd016a44dc2a46:ZXUtd2VzdC0xLmF3cy5mb3VuZC5pbyQxZDdkOWEyMThjYWU0MmUzOGUyZGM3YmJkZDM1ZDA0NyRjZDEwMmZiYTFjNDg0YjBhYTZkYjE1MjRkM2U1NDMzMw==
    @classmethod
    async def connect(cls: type[AsyncElasticService]) -> None:
        cls.client = AsyncElasticsearch(
            cloud_id="Vector_Search_Cluster:ZXVyb3BlLXdlc3QyLmdjcC5lbGFzdGljLWNsb3VkLmNvbTo0NDMkOTRhMjRlMWZhMGY0NDQzNGI3N2Q1MDlkMzU2OTNhNDgkNGM2YmI2MGZmOGI3NGVmYmI1Y2VjZGUxMDA4ZmFiNjA=",
            basic_auth=("elastic", "X2t97nvmIsTRHhxKtVI7o63X")
        )

    @classmethod
    async def close_connection(cls: type[AsyncElasticService]) -> None:
        await cls.client.close()

    @classmethod
    async def search(cls, query: str) -> dict[str, list[dict[str, str]]]:
        product_code_pattern = r'^[A-Z]{2}-\d{3}-\d{2,3}\s?[A-Z]{0,2}$'
        query = query.strip()
        if re.match(product_code_pattern, query):
            filter_query = {
                "query": {
                    "match": {
                        "product_key": query
                    }
                }
            }
        else:
            number_pattern = r'\b(?<!-)(?<!/)\d+(?:\.\d+)?\b(?!\w)(?![/])'
            numbers = [float(num.replace(',', '.')) for num in re.findall(number_pattern, query)]
            query_without_number = re.sub(number_pattern, '', query).strip()

            filter_query = {
                "knn": {
                    "field": "search_vector",
                    "query_vector": get_embedding(query),
                    "k": 10,
                    "num_candidates": 100,
                },
                "query": {
                    "bool": {
                        "must": [
                            {
                                "bool": {
                                    "should": [
                                        {
                                            "match_phrase": {
                                                "description": {
                                                    "query": query,
                                                    "boost": 1.1
                                                },
                                            },
                                        },
                                        {
                                            "match_phrase": {
                                                "main_description_part": {
                                                    "query": query,
                                                    "boost": 1.5
                                                },
                                            },
                                        },
                                        {
                                            "term": {
                                                "description": {
                                                    "value": query,
                                                    "boost": 3
                                                },
                                            }
                                        },
                                        {
                                            "term": {
                                                "search_patterns": {
                                                    "value": query,
                                                    "boost": 1.5
                                                },
                                            }
                                        },
                                        {
                                            "multi_match": {
                                                "query": query_without_number,
                                                "fields": ["description", "search_patterns", "main_description_part"],
                                                "fuzziness": "1"
                                            }
                                        },
                                        {
                                            "multi_match": {
                                                "query": query_without_number,
                                                "fields": ["description", "search_patterns", "main_description_part"],
                                                "boost": 2.5
                                            }
                                        },
                                        {
                                            "term": {
                                                "product_key": {
                                                    "value": query,
                                                    "boost": 10
                                                }
                                            }
                                        },
                                        {
                                            "term": {
                                                "product_key.keyword": {
                                                    "value": query,
                                                    "boost": 20
                                                }
                                            }
                                        },
                                        {
                                            "wildcard": {
                                                "normalized_description": {
                                                    "value": f"*{query_without_number.lower()}*"
                                                }
                                            }
                                        }
                                    ]
                                }
                            }
                        ],
                    }
                }
            }

            if numbers and not re.match(product_code_pattern, query):
                filter_query['query']['bool']['must'].extend(
                    [
                        {
                            "bool": {
                                "must": [
                                    {
                                        "terms": {
                                            "dimensions": numbers
                                        }
                                    }
                                ]
                            }
                        }
                    ]
                )
        # filter_query['sort'] = [
        #     {
        #         "_score": {"order": "desc"}
        #     },
        #     {
        #         "product_key.keyword": {"order": "desc"}
        #     }
        # ]
        print(filter_query)
        result = await AsyncElasticService.client.search(
            index=ES_INDEX_NAME,
            body=filter_query,
            min_score=1
        )
        # filtered_results = []
        # if result['hits']['hits']:
        #     ten_percent_from_top = result['hits']['hits'][0]['_score'] * 0.1
        #     filtered_results = [
        #         res for res in result['hits']['hits'] if res['_score'] >= ten_percent_from_top
        #     ]
        return {
            'search_results': [
                {
                    'product_key': item['_source']['product_key'],
                    'description': item['_source']['description'],
                    'search_patterns': item['_source']['search_patterns'],
                    'score': item['_score']
                }
                for item in result['hits']['hits']
            ]
        }

    @staticmethod
    async def simply_iterate(products_list):
        for product in products_list:
            yield product

    @classmethod
    async def ingest_data_into_es_index(cls) -> None:
        '''Ingest data into ElasticSearch'''
        await cls.excel_to_json()
        await cls.json_to_csv()

        if not (await cls.client.indices.exists(index=cls.es_index_name)):
            await cls.client.indices.create(index=cls.es_index_name, mappings=cls.products_es_model)

        df = pd.read_csv(cls.products_info_csv_filename)

        # Add new field with vector for each product (combines all 'ai_search_patterns' encoded together)
        df['search_vector'] = df['search_patterns'].apply(lambda x: get_embedding(x))

        products_list = df.to_dict(orient='records')
        # Upload all products info to ElasticSearch index
        async for _ in async_streaming_bulk(
            client=cls.client,
            index=cls.es_index_name,
            actions=cls.simply_iterate(products_list)
        ):
            pass

    @classmethod
    async def json_to_csv(cls) -> None:
        # Load data into memory from json file
        with open(cls.products_info_json_filename, 'r', encoding='utf-8') as f:
            data = json.loads(f.read())
        # Normalize data
        df = pd.json_normalize(data)
        # Store data in csv file
        df.to_csv(cls.products_info_csv_filename, index=False, encoding='utf-8')

    @classmethod
    async def excel_to_json(cls) -> None:
        from openpyxl import load_workbook
        print('hey')
        wb1 = load_workbook(filename=cls.inbound_data_excel_files['products_descriptions'])
        sheet_1 = wb1['Sheet1']
        translation_table = str.maketrans('', '', string.whitespace + string.punctuation)
        all_products: dict[str, dict[str, str|list[str]]] = {}

        # Attention: 'max_col' and 'max_row' are the max numbers of col/raws with data on the sheet
        for row in sheet_1.iter_rows(min_row=2, max_col=2, max_row=14525):
            print('sheet 1 row done')
            product_key = str(row[0].value)
            description = row[1].value
            hyphen_desc = description.split(' - ')
            slash_split_desc = description.split(' / ')
            splitted_desc = hyphen_desc or slash_split_desc
            all_products[product_key] = {
                'product_key': product_key,
                'description': description,
                'search_patterns': [description],
                'normalized_description': description.translate(translation_table),
            }
            if len(splitted_desc) >= 2:
                all_products[product_key]['main_description_part'] = splitted_desc[0]
        wb2 = load_workbook(filename=cls.inbound_data_excel_files['ai_search_patterns'])
        sheet_2 = wb2['Sheet1']

        # Attention: 'max_col' and 'max_row' are the max numbers of col/raws with data on the shee
        for row in sheet_2.iter_rows(min_row=2, max_col=2, max_row=5763):
            product_key: str = str(row[1].value)
            ai_search_pattern: str = str(row[0].value)
            iterable_sequence: list[str] = [ai_search_pattern, product_key]
            # case when there are multiple keys in the 'product_key' cell - start -
            if '+' in product_key:
                product_keys = row[1].value.split(' + ')
                for pk in product_keys:
                    try:
                        if all_products[pk].get('search_patterns') is not None:
                            if ai_search_pattern not in all_products[pk]['search_patterns']:
                                all_products[pk]['search_patterns'].append(ai_search_pattern)
                        else:
                            all_products[pk]['search_patterns'] = [ai_search_pattern]
                    except KeyError:
                        for product in all_products:
                            if product.startswith(product_key):
                                if all_products[product].get('search_patterns') is not None:
                                    for value in iterable_sequence:
                                        if value not in all_products[pk]['search_patterns']:
                                            all_products[product]['search_patterns'].append(value)
                                else:
                                    all_products[product]['search_patterns'] = iterable_sequence
                continue
            # case when there are multiple keys in the cell - end -

            try:
                if all_products[product_key].get('search_patterns') is not None:
                    if ai_search_pattern not in all_products[product_key]['search_patterns']:
                        all_products[product_key]['search_patterns'].append(ai_search_pattern)
                else:
                    all_products[product_key]['search_patterns'] = [ai_search_pattern]
            except KeyError:
                for product in all_products:
                    if product.startswith(product_key):
                        if all_products[product].get('search_patterns') is not None:
                            for item in iterable_sequence:
                                if item not in all_products[product]['search_patterns']:
                                    all_products[product]['search_patterns'].append(str(item))

                        else:
                            all_products[product]['search_patterns'] = iterable_sequence

        # Create list of all products info
        pattern = r"[-+]?(?:\d*\.*\d+)"
        all_products_info_list = [v for _, v in all_products.items()]
        for prod in all_products_info_list:
            prod['normalized_search_patterns'] = ''.join(prod['search_patterns']).translate(translation_table)
            numbers = {*[float(num.replace(',', '.')) for num in re.findall(pattern, ''.join(prod['search_patterns']))]}
            numbers.union({*[float(num.replace(',', '.')) for num in re.findall(pattern, prod['description'])]})
            prod['dimensions'] = list(numbers)

        # Store products info in json file
        with open(cls.products_info_json_filename, 'w') as f:
            json.dump(all_products_info_list, f)




AsyncElasticService.connect()
AsyncElasticService.ingest_data_into_es_index()
